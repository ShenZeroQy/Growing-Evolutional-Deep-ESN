

'''''''''''''''''''''
version 2023 03 22
Special for ATP
    @time analysis
    @dynamical update
explaination:
    Growing layer by layer deepESN
    @ X_dim dynamic to a list

    @ average record start and end

    @ solve GUI_sig_and_err bug
    
    @ estimate spectural radious 

    @AMI estimation Shhanno Entropy


version 2022 11 25
explaination:
    generate high efficiency deepESN
@1 X_dim veclized
    dynamic stack 
    dynemic physical storage 

@alternative GPU
    by Pre_def_USE_GPU

@merge method veclized 
    always merge top layer

@growing layer by layer
'''''''''

#from matplotlib import projections


Pre_def_USE_GPU=1
Pre_def_PSO=1
Pre_def_Anayis=0
from numpy import *
# import math
import matplotlib.pyplot as plt
import time 
from openpyxl import Workbook,load_workbook
import datetime
import os 
import numba
import HisgEntropy as Shhanno
if(Pre_def_USE_GPU):
    import CudaEntropy as Entropy

else:
    import LocalEntropy as Entropy
if(Pre_def_PSO):
    from sko.PSO import PSO
else:
    pass
# Dictlized Prams
# predefed Pram
Default_Pram={
        'Exp_dada_file':'./data/ForAPT/Dtrac0301172756.txt',
        
        'Pram_galaph' : 0.92,

        'Pram_U_dim_K' : 2,
        'Pram_Stack':6,
        'Pram_X_dim':[20,20,20,20,20,20,20,20,20,20,20,20,20,20,20,20],
        # 'Pram_X_dim':[40,40,40,40,40,40,40,40,40,40,40,40,40,40,40,40],
        'Pram_X_add':10,
        
        'Pram_Y_dim_L': 1,#muliti 2X 
        'Pram_Y_delay_step': 4,   
        'Pram_initLen' : 100,
        'Pram_trainLen' : 200,
        'Pram_testLen':220,
        'Pram_trainLenMAX' : 2000,#limit max trainlen
        }
Default_Oram={
        'spare_rate':1.0,
        'ampWi':0.847363,
        'ampWp':0.49497293,
        'ampWr':0.8,
        'reg_fac':1e-9,
    }



class EchoStateNetwork:
    #all variables storaged in CUP
    def __init__(self,Pram,Oram):
        self.Y_delay_step=Pram['Pram_Y_delay_step']
        self.U_dim=Pram['Pram_U_dim_K']
        self.Y_dim=Pram['Pram_Y_dim_L']
        # self.X_dim=Pram['Pram_X_min']       
        # self.GroupPool_num=Pram[ 'Groupnum']
        self.galaph=Pram['Pram_galaph']
        self.initLen=Pram['Pram_initLen']
        self.testLen=Pram['Pram_testLen']
        self.trainLen=Pram['Pram_trainLen']
        #load ESN Pram
        self.ampWi=Oram['ampWi']
        self.ampWc=Oram['ampWp']
        self.ampWr=Oram['ampWr']
        self.Reg_fac=Oram['reg_fac']
        self.SpareRate=Oram['spare_rate']
        self.InitX_dim=Pram['Pram_X_add']
        # self.ReserveX_dim=Pram['Pram_X_min']
        self.Stacklayer=Pram['Pram_Stack']
        self.Stack=1
        self.X_dim=list()
        #modefy offest to shift
        name=Pram['Exp_dada_file']
        dataX,dataY=self.loadData(name,offset=1)
        U_in,Y_out=self.H_EncodeData(dataX,dataY, 5000)
        self.U_in=U_in
        self.Y_out=Y_out
        self.U_init=U_in[:,0: self.initLen]
        self.U_train=U_in[:, self.initLen: self.trainLen]
        self.Y_train=Y_out[:, self.initLen: self.trainLen]
        self.U_tt=U_in[:, self.initLen: self.testLen]
        self.Y_tt=Y_out[:, self.initLen: self.testLen]
        self.TrainProcessLen= self.trainLen- self.initLen
        #storage mats
        self.offset=0
        #self.Inilize_reservoir()
        #self.Init_reservior(U_init)
        #self.TrainProcessX=self.Train_reservoir(U_train,Y_train)
        return
    def Update(self,offset):
        start_time=time.time()
        end_time=time.time()
        U_in=self.U_in
        Y_out=self.Y_out
        self.trainLen=self.trainLen+offset
        self.testLen=self.testLen+offset
        self.U_train=U_in[:, self.initLen: self.trainLen]
        self.Y_train=Y_out[:, self.initLen: self.trainLen]
        self.U_tt=U_in[:,self.initLen:self.testLen]
        self.Y_tt=Y_out[:,self.initLen:self.testLen]
        self.TrainProcessLen= self.trainLen- self.initLen
        while (end_time-start_time<0.01):
            end_time=time.time()
        return
    def loadData(self,name,offset:int):
        data=loadtxt(name)
        aveg=mean(data)
        rOram=amax(data)-amin(data)
        # data=(data-aveg)/rOram
        # data=(data-aveg)
        #data=data[offset:offset+5000]
        Xaxis=data[:,0]
        Yaxis=data[:,1]
        return Xaxis,Yaxis
    # @numba.jit
    def H_EncodeData(self,dataX,dataY,length:int):
        U_in=zeros((length,self.U_dim*2))
        #tensor_U_in=torch.from_numpy(U_in)
        Y_out=zeros((length,self.Y_dim*2))
        for i in range(self.U_dim):
            U_in[:,i]=dataX[i:i+length]
            U_in[:,self.U_dim+i]=dataY[i:i+length]
           
        for i in range(self.Y_dim):
            Y_out[:,i]=dataX[i+self.U_dim+1+self.Y_delay_step : i+self.U_dim+length+1+self.Y_delay_step]           
            Y_out[:,self.Y_dim+i]=dataY[i+self.U_dim+1+self.Y_delay_step : i+self.U_dim+length+1+self.Y_delay_step]   
        self.U_dim=self.U_dim*2
        self.Y_dim=self.Y_dim*2
        return U_in.T, Y_out.T
    # @numba.jit
    def UspanX(self,uj,alaph):
        #this function is essential asure all varables are vertical vectors 
        self.GroupX[0] = (1 - alaph) * self.GroupX[0] + alaph *tanh(dot(self.GroupWin[0], uj)+dot(self.GroupW[0], self.GroupX[0]))
        for i in range(1,self.Stack):
            self.GroupX[i] = (1 - alaph) * self.GroupX[i] + alaph * tanh(dot(self.GroupC[i-1], self.GroupX[i-1])+dot(self.GroupW[i], self.GroupX[i])) 

        U_X = self.GroupX[0]
        for i in range(1,self.Stack):        
            U_X = concatenate((U_X, self.GroupX[i]), axis=0)
        return U_X
    def Inilize_Stack_a_reservoir(self,X_dim):
        #inilize a layer assign given number of neurons
        #initial zero state of X 
        #xj = zeros((X_dim, 1))
        xj = random.rand(X_dim, 1)

        self.GroupX.append(xj)
       
        #initial  transection of C           
        ci = random.rand(X_dim,self.X_dim[self.Stack-1])-0.5
        ci=ci*self.ampWc
        self.GroupC.append(ci)

        # #initial random W by scaleing lamda
        PCAcore= random.rand(X_dim,X_dim)-0.5
        PCAspare=self.H_Sparelize_to01(PCAcore,X_dim,X_dim,self.SpareRate)
        PCAcore= random.rand(X_dim,X_dim )-0.5
        PCAcore= PCAcore*PCAspare
        e_vals,e_vecs = linalg.eig(PCAcore)
        Lamda =abs(e_vals).max()
        Wi=PCAcore/Lamda*self.ampWr            
        self.GroupW.append(Wi)

        #external
        self.X_dim.append(X_dim)
        self.ExistNode=concatenate((self.ExistNode,ones((X_dim,)).astype(int32)),axis=0)
        self.Stack=self.Stack+1

    def Inilize_First_reservoir(self,X_dim):
        #inilize a layer assign given number of neurons

        #initial zero state of X
        self.GroupX = list()  
        #xj = zeros((X_dim, 1))
        xj = random.rand(X_dim, 1)

        self.GroupX.append(xj)
       
        #initial random Win
        self.GroupWin = list()
        Wini = random.rand(X_dim,self.U_dim )-0.5
        Wini=Wini*self.ampWi
        self.GroupWin.append(Wini) 

        #initial  empty transection of C
        self.GroupC = list()
        # for i in range(1,self.GroupPool_num):           
        #     ci = random.rand(self.X_dim,self.X_dim)-0.5
        #     ci=ci*self.ampWc
        #     self.GroupC.append(ci)

        #initial  stable  W
        self.GroupW = list()
        # #initial random W by scaleing lamda
        PCAcore= random.rand(X_dim,X_dim)-0.5
        PCAspare=self.H_Sparelize_to01(PCAcore,X_dim,X_dim,self.SpareRate)
        PCAcore= random.rand(X_dim,X_dim )-0.5
        PCAcore= PCAcore*PCAspare
        e_vals,e_vecs = linalg.eig(PCAcore)
        Lamda =abs(e_vals).max()
        Wi=PCAcore/Lamda*self.ampWr            
        self.GroupW.append(Wi)

        #external

        self.X_dim.append(X_dim)
        self.ExistNode=ones((X_dim,)).astype(int32)
        self.Stack=1
        
        return
      

    def Init_reservior(self,U_init,alaph=1):
        # for i in range(self.Stack):
        #     self.GroupX[i]=zeros((self.X_dim,1))
        
        initLen=U_init.shape[1] 
        for j in range(initLen):
            uj=U_init[:,j:j+1]#input
            #uj=uj.reshape((self.U_dim,1))
            #uj = self.initU[j]
            self.UspanX(uj,alaph)        
        #record X0 after span
        self.InitialX=list()
        for i in range(self.Stack):
            self.InitialX.append(self.GroupX[i])
        return
    def Reinit_reservoir(self):
        #restore default X states
        for i in range(self.Stack):
            self.GroupX[i]=self.InitialX[i]
        return
    def Clear_reservoir(self):
        #set X states to zero
        for i in range(self.GroupPool_num):
            self.GroupX[i]=zeros((self.X_dim[i],1))

        return
    
    @numba.jit
    def Train_reservoir(self,U_train,Y_train):    
        alaph=self.galaph
        X_train=zeros((sum(self.X_dim),U_train.shape[1]))
        for i in range(U_train.shape[1]):
            X_train[:,i:i+1]=self.UspanX(U_train[:,i:i+1],alaph)
 
        #once solution
        #Wout = dot(Y_trian, linalg.pinv(X_train))
        
        self.PWout=linalg.inv(dot(X_train, X_train.T) + self.Reg_fac * eye(X_train.shape[0]))
        self.Wout = dot(self.PWout,dot(X_train,Y_train.T ) ).T 
        
        #iterater solution learning
        # self.OnceWout=self.Wout
        # self.PWout=self.PWout+ self.Reg_fac * eye(shape(self.PWout)[0])
        return X_train
    def Verify_Train_reservoir(self, train_num,savename):
        y1=zeros((self.Y_dim,train_num))
        y2=zeros((self.Y_dim,train_num))
        self.errt=list()
        for j in range(train_num):
            uj = self.trainU[j]
            U_Xj = self.trainX[j]
            yvj=dot(self.Wout,U_Xj)
            for i in range(self.Y_dim):
                y1[i][j]=yvj[i]
                y2[i][j]=self.trainY[j][i]
            #record error
            self.errt.append((yvj-self.trainY[j])**2)    
        mse=sum(self.errt)/train_num
         
        plt.figure(1)
        plt.plot(range(train_num), y1[0,:], ls="dotted", lw=0.3,color='b' ,label='yo')
        plt.plot(range(train_num), y2[0,:], ls="-", lw=0.1,color='r', label='Ys')  
        plt.title('train signal and train result') 
        plt.legend(loc = 'upper right')        
        plt.savefig(savename)
        plt.close(plt.figure(1))
        mse=sum(self.errt)/train_num
        return mse
    
    # @numba.jit
    def Validate_test_data_constant(self, U_test):     
        # make sure U_test is constant reinit if necessary
        alaph=self.galaph
        X_test=zeros((sum(self.X_dim),U_test.shape[1]))
        for i in range(U_test.shape[1]):
            X_test[:,i:i+1]=self.UspanX(U_test[:,i:i+1],alaph)
        Y_test = dot(self.Wout, X_test)
        #X_test used for analysis
        return Y_test ,X_test
    # @numba.jit
    def Validate_test_data_once(self, U_test):     
        # make sure U_test is constant reinit if necessary
        alaph=self.galaph
        X_test=zeros((sum(self.X_dim),self.testLen-self.trainLen))
        for i in range(-self.trainLen+self.testLen):
            
            X_test[:,i:i+1]=self.UspanX(U_test[:,self.trainLen-self.initLen+i:self.trainLen-self.initLen+i+1],alaph)
        Y_test = dot(self.Wout, X_test)
        #X_test used for analysis
        return Y_test ,X_test
   
    def H_Sparelize_to01(self,Wi,x:int,y:int,SpareRate:float):
        z=x*y
        zr=z-int(z*SpareRate)
        Wi=Wi.reshape(z,)
        index=argsort(Wi)
        index0=index[0:zr]
        index1=index[zr:]
        for i in index0:
            Wi[i]=0
        for i in index1:
            Wi[i]=1
        Wi=Wi.reshape(x,y)
        return Wi
  
    def H_Physical_delect(self,arr:ndarray,r:int,c:int):
        if(c>=0):
            arr[:, c] = arr[:, -1]
        if(r>=0):    
            arr[r,:] = arr[-1,:]
        if(r>=0 and c>=0):    
            arr2 = arr[:-1, :-1]
        else:
            if(r>=0):
                arr2 = arr[:-1, :]
            if(c>=0):
                arr2 = arr[:, :-1]
        return arr2
    
    def H_Physical_delect_anode(self,lev,ind):
        #swap with last one
        #reservoir state
        self.GroupX[lev]=self.H_Physical_delect(self.GroupX[lev],ind,-1)

        #input terminal
        if(lev==0):
            self.GroupWin[0]=self.H_Physical_delect(self.GroupWin[0],ind,-1)
        else:
            self.GroupC[lev-1]=self.H_Physical_delect(self.GroupC[lev-1],ind,-1)
        
        #squire span
        self.GroupW[lev]=self.H_Physical_delect(self.GroupW[lev],ind,ind)

        #output terminal
        if(lev>=self.Stack-1):
            pass
        else:
            self.GroupC[lev]=self.H_Physical_delect(self.GroupC[lev],-1,ind)

        self.X_dim[lev]=self.X_dim[lev]-1
        return
    def CCN_MergeQ2(self,lev:int,indi:int,indj:int):
           
        # self.ExistNode[Nodej]=0

        #Win[lev]indi <-indj *0.5
        #Wres[lev]in indi-<indj *0.5
        #Wres[lev]out  indj->indi *0.5
        #Wc[lev]indj-> set0


        #CHECK a,b
        wii=self.GroupW[lev][indi][indi]
        wij=self.GroupW[lev][indi][indj]
        wji=self.GroupW[lev][indj][indi]
        wjj=self.GroupW[lev][indj][indj]
        det=(wjj-wii)**2+4*wji*wij
        a=0.5
        b=0.5
        p=0.5*(wii+wij+wji+wjj)
        p2=0
        if(wij==0):
            if(wii==wjj):
                pass
            else:
                # linear solution
                rateab=wji/(wjj-wii)
                b=1/(rateab+1)
                a=1-b
                p=rateab*wij+wjj
                p2=wii+wji/rateab-p
        else:
            if(det>=0):
                ra=(sqrt(det)-wjj+wii)/2/wij
                rb=(-sqrt(det)-wjj+wii)/2/wij
                rateab=min(ra,rb)
                b=1/(rateab+1)
                a=1-b
                p=rateab*wij+wjj
                p2=wii+wji/rateab-p
        if(p2>1e-5):
            print('det p='+str(p2))
        #input terminal：
        #按行horazon处理
        if(lev==0):
        #lev 0
            for i in range(self.U_dim):
                self.GroupWin[lev][indi][i]=(a*self.GroupWin[lev][indi][i]+b*self.GroupWin[lev][indj][i])           
                self.GroupWin[lev][indj][i]=0
        else:
            for i in range(self.X_dim):
                self.GroupC[lev-1][indi][i]=(a*self.GroupC[lev-1][indi][i]+b*self.GroupC[lev-1][indj][i])
                self.GroupC[lev-1][indj][i]=0  

        #output terminal：
        #按列vertical处理
        if(self.GroupPool_num-1==lev):
            #last lev
            pass
        else:
            for i in range(self.X_dim):
                #self.GroupC[lev][i][indi]=(self.GroupC[lev][i][indi]+self.GroupC[lev][i][indj])/2
                self.GroupC[lev][i][indi]=(self.GroupC[lev][i][indi]+self.GroupC[lev][i][indj])
                self.GroupC[lev][i][indj]=0  
        #span matriax:
        #both
        for i in range(self.X_dim):
            self.GroupW[lev][indi][i]=(a*self.GroupW[lev][indi][i]+b*self.GroupW[lev][indj][i])
            self.GroupW[lev][indj][i]=0 
            #self.GroupW[lev][i][indi]=(self.GroupW[lev][i][indi]+self.GroupW[lev][i][indj])/2 
            self.GroupW[lev][i][indi]=(self.GroupW[lev][i][indi]+self.GroupW[lev][i][indj])
            self.GroupW[lev][i][indj]=0  
        self.GroupW[lev][indi][indi]=p
        self.GroupW[lev][indj][indj]=0   

        return
    
    def CCN_Merge_Top(self,lev,indi:int,indj:int,Q2=0,logic=1):
        #Merge lndi and indj in lev
        # self.ExistNode[Nodej]=0
        #logic:logic delect(1) or physical delect(0)?
        
        #Win[lev][indi,:] <-indj *0.5
        #Wres[lev]in indi-<indj *0.5
        #Wres[lev]out  indj->indi *0.5
        #Wc[lev]indj-> set0

        #determine weight
        wii=self.GroupW[lev][indi][indi]
        wij=self.GroupW[lev][indi][indj]
        wji=self.GroupW[lev][indj][indi]
        wjj=self.GroupW[lev][indj][indj]
        a=0.5
        b=0.5
        p=0.5*(wii+wij+wji+wjj)
        if(Q2):
            p2=0
            #modefy a,b,p
            det=(wjj-wii)**2+4*wji*wij
            if(wij==0):
                if(wii==wjj):
                    pass
                else:
                    # linear solution
                    rateab=wji/(wjj-wii)
                    b=1/(rateab+1)
                    a=1-b
                    p=rateab*wij+wjj
                    p2=wii+wji/rateab-p
            else:
                if(det>=0):
                    ra=(sqrt(det)-wjj+wii)/2/wij
                    rb=(-sqrt(det)-wjj+wii)/2/wij
                    rateab=min(ra,rb)
                    b=1/(rateab+1)
                    a=1-b
                    p=rateab*wij+wjj
                    p2=wii+wji/rateab-p
            if(p2>1e-5):
                print('det p='+str(p2))
            #done Q2

            #input terminal：
        #按行horazon处理
        if(lev==0):
        #lev 0
           for i in range(self.U_dim):
                self.GroupWin[lev][indi][i]=(a*self.GroupWin[lev][indi][i]+b*self.GroupWin[lev][indj][i])           
                if(logic):
                    self.GroupWin[lev][indj][i]=0
        else:
            for i in range(self.X_dim[lev-1]):
                self.GroupC[lev-1][indi][i]=(a*self.GroupC[lev-1][indi][i]+b*self.GroupC[lev-1][indj][i])
                if(logic):
                    self.GroupC[lev-1][indj][i]=0    
                

        #output terminal：
        #按列vertical处理
        if(self.Stack-1==lev):
            #last lev
            pass
        else:
            for i in range(self.X_dim[lev+1]):
                #self.GroupC[lev][i][indi]=(self.GroupC[lev][i][indi]+self.GroupC[lev][i][indj])/2
                self.GroupC[lev][i][indi]=(self.GroupC[lev][i][indi]+self.GroupC[lev][i][indj])
                if(logic):
                    self.GroupC[lev][i][indj]=0  
        #span matriax:
        #both
        for i in range(self.X_dim[lev]):
            self.GroupW[lev][indi][i]=a*self.GroupW[lev][indi][i]+b*self.GroupW[lev][indj][i]
            if(logic):
                self.GroupW[lev][indj][i]=0 
        for i in range(self.X_dim[lev]):
            self.GroupW[lev][i][indi]=(self.GroupW[lev][i][indi]+self.GroupW[lev][i][indj])
            if(logic):
                self.GroupW[lev][i][indj]=0  
        self.GroupW[lev][indi][indi]=p
        self.GroupW[lev][indj][indj]=0
        #storage 
        if(logic):  
            sumpre=sum(self.X_dim)-self.X_dim[self.Stack-1]     
            self.ExistNode[sumpre+indj]=0
        else:
            self.H_Physical_delect_anode(lev,indj)
        return
    def CCN_Cut(self,lev:int,ind:int,logic=1):
        #given X_dim
        #given  GroupNum

        # self.ExistNode[Nodei]=0

        #Win[lev]->ind set0
        #Wres[lev]->ind set0
        #Wres[lev]ind-> set0
        #Wc[lev]ind-> set0
        #Wc[lev-1]->ind set0
        
        if(logic):  
            #input terminal：
            if(lev==0):
                #lev 0
                for i in range(self.U_dim):
                    self.GroupWin[lev][ind][i]=0            
            else:
                for i in range(self.X_dim[lev-1]):
                    self.GroupC[lev-1][ind][i]=0  

            #output terminal：
            if(self.Stack-1==lev):
                #last lev
                pass
            else:
                for i in range(self.X_dim[lev+1]):
                    self.GroupC[lev][i][ind]=0  
            #span matriax:
            for i in range(self.X_dim[lev]):
                self.GroupW[lev][ind][i]=0  
                self.GroupW[lev][i][ind]=0  
            sumpre=sum(self.X_dim)-self.X_dim[self.Stack-1]     
            self.ExistNode[sumpre+ind]=0
        else:
            self.H_Physical_delect_anode(lev,ind)
        return

    def GUIrecord(self,fname='./cache/GUIrecord.txt'):
        with open(fname,'w+') as fp:
            fp.write('Generated at:'+str(datetime.datetime.now())+'\n')
            fp.write('---------------------------\n')
            fp.write('ESNstack'+str(self.Stack)+'\n')
            fp.write('U_dim:'+str(self.U_dim)+'\n')
            fp.write('X_dim:'+str(self.X_dim)+'\n')
            fp.write('---------------------------\n')
            fp.write('Win:\n')
            fp.write('__lev0:\n')
            for line in self.GroupWin[0]:
                    fp.write(str(line))
                    fp.write('\n')
            fp.write('Wp:\n')
            for i in range(self.Stack-1):
                fp.write('__lev'+str(i+1)+':\n')
                for line in self.GroupC[i]:
                    fp.write(str(line))
                    fp.write('\n')
            fp.write('---------------------------\n')
            fp.write('Wr:\n')
            for i in range(self.Stack):
                fp.write('__lev'+str(i)+':\n')
                for line in self.GroupW[i]:
                    fp.write(str(line))
                    fp.write('\n')
            fp.write('---------------------------\n')
            fp.write('StateX\n')
            for i in range(self.Stack):
                fp.write('__lev'+str(i)+':\n')
                for line in self.GroupX[i]:
                    fp.write(str(line))
                    fp.write('  ')
                fp.write('\n')


class ESNsimulater:
    def __init__(self,Pram,Oram):
        self.Pram=Pram
        self.Oram=Oram
        self.DictRes={
            'NRMSEtrain_aveg':0.0,
            'NRMSEtrain_std':0.0,
            'NRMSEtest_aveg':0.0,
            'NRMSEtest_std':0.0,
            # 'RMSEtrain_aveg':0.0,
            # 'RMSEtrain_std':0.0,
            # 'RMSEtest_aveg':0.0,
            # 'RMSEtest_std':0.0,
            'EntropyX_aveg':0.0,
            'EntropyX_std':0.0,
        }

    def Increasingly_evaluateMergeHE(self,ldir='./default/',method=1,Q2=0):
        #load pram
        endpoint=6000
        inc_scale=1.5
        X_dim_stack=self.Pram['Pram_X_dim']
        X_dim_add=self.Pram['Pram_X_add']
        Stacklayer=self.Pram['Pram_Stack']
        Inc_MaxTrainLen=self.Pram['Pram_trainLenMAX']
        Inc_out=0
        #method 0：distance
        #fix 0
        self.Oram['Method']=method
        self.Oram['Q2']=Q2
        #self.EP=EP
        #prepare output fig and files
        self.mkdir(ldir)
        self.ExcelRecord(ldir+"pram")
        logfilepointer=open(ldir+'log.txt','w')
        # simulate ESN
        ESN_0 = EchoStateNetwork(Pram=self.Pram,Oram=self.Oram)
        
        #allcate layer storage
        eva_NRMSE_test=zeros(( Stacklayer*2,))
        eva_NRMSE_train=zeros(( Stacklayer*2,))
        eva_ASE=zeros(( Stacklayer*2,))
        eva_AMI=zeros(( Stacklayer*2,))
        #ADD layer by layer
        Inc_Trueout=zeros((ESN_0.Y_dim,endpoint))
        #first layer
        dirl0=ldir+'layer0/'
        self.mkdir(dirl0)
        logfilepointer.write('add first layer neuron:'+str(X_dim_add+X_dim_stack[0])+'\n')
        ESN_0.Inilize_First_reservoir(X_dim_add+X_dim_stack[0])
        ESN_0.GUIrecord(dirl0+'Bef.txt')
        Merge_Max_Similarity=zeros((X_dim_add,))
        Merge_Nrmse_test=zeros((X_dim_add,))
        Merge_Nrmse_train=zeros((X_dim_add,))
        Merge_ASE=zeros((X_dim_add,))
        Merge_AMI=zeros((X_dim_add,))
        Merge_Lamda=zeros((X_dim_add,))
        c=0
        
        inc_sumoffset=0
        while(ESN_0.X_dim[0]>X_dim_stack[0]):
            #run
            ESN_0.Init_reservior(ESN_0.U_init)
            traintime=time.time()
            TrainProcessX=ESN_0.Train_reservoir(ESN_0.U_train,ESN_0.Y_train)
            traintime=time.time()-traintime
            print('firstLayer,Traing with:'+str(traintime))            
            X_num,T_num= TrainProcessX.shape
            #estimate Similarity

            opttime1=time.time()
            SM=Entropy.SMEstimater(TrainProcessX,X_num,T_num,method)
            opttime1=time.time()-opttime1
            
            MaxSM=amax(SM)
            winner=argwhere(SM== MaxSM)
            mergei,mergej=winner[0]
            Merge_Max_Similarity[c]=MaxSM
           
            
            #real time learn
            learntime=time.time()
            learnedYout,learnedValidateX=ESN_0.Validate_test_data_once(ESN_0.U_tt)
            learntime=time.time()-learntime
            print('firstLayer,Learning with:'+str(learntime))
            Inc_Trueout[:,inc_sumoffset]=mean(learnedYout,axis=1)
           

            #validate
            ESN_0.Reinit_reservoir()
            Yout,ValidateX=ESN_0.Validate_test_data_constant(ESN_0.U_tt)
            #draw fig
            if(c==0):
                self.GUI_sig_and_err(Yout,ESN_0.Y_tt,ESN_0.TrainProcessLen,dirl0,0)
            
            #estimate err
            err=(Yout-ESN_0.Y_tt)**2
            Ynorm=mean(ESN_0.Y_tt)*ones(Yout.shape)
            nerr=(Yout-Ynorm)**2
            MSE_train=sum(err[0:ESN_0.TrainProcessLen])/ESN_0.TrainProcessLen
            NRMSE_train=sqrt(sum(err[0,0:ESN_0.TrainProcessLen])/sum(nerr[0,0:ESN_0.TrainProcessLen]))
            tlen=Yout.shape[1]-ESN_0.TrainProcessLen
            MSE_test=sum(err[0,ESN_0.TrainProcessLen:])/tlen
            NRMSE_test=sqrt(sum(err[0,ESN_0.TrainProcessLen:Yout.shape[1]])/sum(nerr[0,ESN_0.TrainProcessLen:Yout.shape[1]]))
            Merge_Nrmse_train[c]=NRMSE_train
            Merge_Nrmse_test[c]=NRMSE_test

            #estimate SP
            PCAcore=ESN_0.GroupW[0]
            e_vals,e_vecs = linalg.eig(PCAcore)
            Lamda =abs(e_vals).max()
            Merge_Lamda[c]=Lamda

            #estimate entropy
            X_num,T_num= ValidateX.shape 
            #Renyi
            ASE=Entropy.EntropyEstimater(ValidateX,X_num,T_num)
            Merge_ASE[c]=ASE  
            #Shhanno
            AMI=Shhanno.IXY(ValidateX,Yout,X_num,ESN_0.Y_dim,T_num)  
            Merge_AMI[c]=AMI 

            logfilepointer.write('Will merge:'+str([mergei,mergej])+'\n')
            #print('cut_merge:'+str(max([cuti,cutj])))
            
            opttime2=time.time()
            ESN_0.CCN_Merge_Top(0,mergei,mergej,Q2,logic=0) 
            opttime2=time.time()-opttime2+opttime1
            print('firstLayer,optimied with:'+str(opttime2))
            offset=int((opttime2+traintime)/0.001/inc_scale)
            ESN_0.Update(offset)
            inc_sumoffset=inc_sumoffset+offset
            c=c+1
        
        logfilepointer.write('first layer shrinked to:'+str(ESN_0.X_dim[0])+'\n') 
        ESN_0.GUIrecord(dirl0+'Aft.txt')
        # self.GUI_sig_and_err(Yout,ESN_0.Y_tt,ESN_0.TrainProcessLen,dirl0,10)   

        #save to evas[0]
         #start point 
        eva_AMI[0]=Merge_AMI[0]
        eva_ASE[0]=Merge_ASE[0]
        eva_NRMSE_train[0]=Merge_Nrmse_train[0]
        eva_NRMSE_test[0]=Merge_Nrmse_test[0]
         #end point 
        eva_AMI[1]=Merge_AMI[-1]
        eva_ASE[1]=Merge_ASE[-1]
        eva_NRMSE_train[1]=Merge_Nrmse_train[-1]
        eva_NRMSE_test[1]=Merge_Nrmse_test[-1]

        #savefiles of first layer
        if(1):
            savetxt(dirl0+'NRMSE_test.txt',Merge_Nrmse_test, fmt='%f', delimiter=',')
            savetxt(dirl0+'NRMSE_train.txt',Merge_Nrmse_train, fmt='%f', delimiter=',')
            savetxt(dirl0+'ASE.txt',Merge_ASE, fmt='%f', delimiter=',')
            savetxt(dirl0+'AMI.txt',Merge_AMI, fmt='%f', delimiter=',')
            savetxt(dirl0+'MAX_corr.txt',Merge_Max_Similarity , fmt='%f', delimiter=',')
            savetxt(dirl0+'Lamda.txt',Merge_Lamda, fmt='%f', delimiter=',')
            #savetxt(0ldir+'CE'+'.txt',eva_Ce[0:Cutsum], fmt='%f', delimiter=',')       
            logfilepointer.write("\nexp data researved\n")
        
        for i in range(1 , Stacklayer):
            #higer layer
            
            if(Inc_out):
                break#end growing
            
            #first add a layer
            dirli=ldir+'layer' +str(i)+'/'
            self.mkdir(dirli)
            logfilepointer.write('add a layer'+str(i)+' of neuron:'+str(X_dim_add+X_dim_stack[i])+'\n')
            X_num_fronter=sum(ESN_0.X_dim)
            growtime=time.time()
            ESN_0.Inilize_Stack_a_reservoir(X_dim_add+X_dim_stack[i])
            growtime=time.time()-growtime
            print('Add a new layer'+str(i)+',grown with:'+str(growtime))
            
            #update data
            offset=int((opttime2+traintime)/0.001/inc_scale)
            ESN_0.Update(offset)
            inc_sumoffset=inc_sumoffset+offset
            
            ESN_0.GUIrecord(dirli+'Bef.txt')
            Merge_Max_Similarity=zeros((X_dim_add,))
            Merge_Nrmse_test=zeros((X_dim_add,))
            Merge_Nrmse_train=zeros((X_dim_add,))
            Merge_ASE=zeros((X_dim_add,))
            Merge_AMI=zeros((X_dim_add,))
            Merge_Lamda=zeros((X_dim_add,))
            c=0

            while(Inc_out==0 and  ESN_0.X_dim[i]>X_dim_stack[i]):
                #run
                ESN_0.Init_reservior(ESN_0.U_init)
                traintime=time.time()
                TrainProcessX=ESN_0.Train_reservoir(ESN_0.U_train,ESN_0.Y_train)
                traintime=time.time()-traintime
                print('layer'+str(i)+',Traing with:'+str(traintime))            
                X_num,T_num= TrainProcessX.shape
                #estimate Similarity
                opttime1=time.time()
                SM=Entropy.SMEstimater(TrainProcessX,X_num,T_num,method)
                opttime1=time.time()-opttime1
                #ban elter neurons 
                SM[0:X_num_fronter,0:X_num_fronter]=zeros((X_num_fronter,X_num_fronter))

                MaxSM=amax(SM)
                winner=argwhere(SM== MaxSM)
                mergei,mergej=winner[0]
                
                Merge_Max_Similarity[c]=MaxSM
                
                #real time learn
                learntime=time.time()
                learnedYout,learnedValidateX=ESN_0.Validate_test_data_once(ESN_0.U_tt)
                learntime=time.time()-learntime
                print('layer'+str(i)+',Learning with:'+str(learntime))
                Inc_Trueout[:,inc_sumoffset]=mean(learnedYout,axis=1)
                
                ESN_0.Reinit_reservoir()
                #validate
                Yout,ValidateX=ESN_0.Validate_test_data_constant(ESN_0.U_tt)
                #draw fig
                if(c==0):
                    self.GUI_sig_and_err(Yout,ESN_0.Y_tt,ESN_0.TrainProcessLen,dirli,0)
                
                #estimate err
                err=(Yout-ESN_0.Y_tt)**2
                Ynorm=mean(ESN_0.Y_tt)*ones(Yout.shape)
                nerr=(Yout-Ynorm)**2
                MSE_train=sum(err[0:ESN_0.TrainProcessLen])/ESN_0.TrainProcessLen
                NRMSE_train=sqrt(sum(err[0,0:ESN_0.TrainProcessLen])/sum(nerr[0,0:ESN_0.TrainProcessLen]))
                tlen=Yout.shape[1]-ESN_0.TrainProcessLen
                MSE_test=sum(err[0,ESN_0.TrainProcessLen:])/tlen
                NRMSE_test=sqrt(sum(err[0,ESN_0.TrainProcessLen:Yout.shape[1]])/sum(nerr[0,ESN_0.TrainProcessLen:Yout.shape[1]]))
                Merge_Nrmse_train[c]=NRMSE_train
                Merge_Nrmse_test[c]=NRMSE_test

                #estimate SP
                PCAcore=ESN_0.GroupW[i]
                e_vals,e_vecs = linalg.eig(PCAcore)
                Lamda =abs(e_vals).max()
                Merge_Lamda[c]=Lamda
                    
                #estimate entropy
                X_num,T_num= ValidateX.shape    
                #Renyi
                ASE=Entropy.EntropyEstimater(ValidateX,X_num,T_num)
                Merge_ASE[c]=ASE  
                #Shhanno
                AMI=Shhanno.IXY(ValidateX,Yout,X_num,ESN_0.Y_dim,T_num)  
                Merge_AMI[c]=AMI 
                logfilepointer.write('Will cut_merge:'+str([mergei,mergej])+'\n')
                #print('cut_merge:'+str(max([cuti,cutj])))
                
                opttime2=time.time()
                if(mergei<X_num_fronter):
                    ESN_0.CCN_Cut(ESN_0.Stack-1,mergej-X_num_fronter,logic=0)
                    logfilepointer.write('Will cut:'+str(mergej)+'\n')
                else:
                    if(mergej<X_num_fronter):
                        ESN_0.CCN_Cut(ESN_0.Stack-1,mergei-X_num_fronter,logic=0)
                        logfilepointer.write('Will cut:'+str(mergei)+'\n')
                    else:
                        ESN_0.CCN_Merge_Top(ESN_0.Stack-1,mergei-X_num_fronter,mergej-X_num_fronter,Q2,logic=0) 
                        logfilepointer.write('Will merge:'+str([mergei,mergej])+'\n')
                
                opttime2=time.time()-opttime2+opttime1
                print('layer'+str(i)+',optimied with:'+str(opttime2))
                offset=int((opttime2+traintime)/0.001/inc_scale)
                ESN_0.Update(offset)
                inc_sumoffset=inc_sumoffset+offset
                if(ESN_0.trainLen>=Inc_MaxTrainLen or learntime>0.0275):
                    Inc_out=1
                c=c+1
                
            logfilepointer.write('layer'+str(i)+' shrinked to:'+str(X_dim_stack[i])+'\n') 
            ESN_0.GUIrecord(dirli+'Aft.txt')
            # self.GUI_sig_and_err(Yout,ESN_0.Y_tt,ESN_0.TrainProcessLen,dirli,i)   
            
            #save to eva s
                #start ppoint 
            eva_AMI[2*i]=Merge_AMI[0]
            eva_ASE[2*i]=Merge_ASE[0]
            eva_NRMSE_train[2*i]=Merge_Nrmse_train[0]
            eva_NRMSE_test[2*i]=Merge_Nrmse_test[0]
                #end point 
            eva_AMI[2*i+1]=Merge_AMI[-1]
            eva_ASE[2*i+1]=Merge_ASE[-1]
            eva_NRMSE_train[2*i+1]=Merge_Nrmse_train[-1]
            eva_NRMSE_test[2*i+1]=Merge_Nrmse_test[-1]
            #savefiles of  layer i
            if(1):
                savetxt(dirli+'NRMSE_test.txt',Merge_Nrmse_test, fmt='%f', delimiter=',')
                savetxt(dirli+'NRMSE_train.txt',Merge_Nrmse_train, fmt='%f', delimiter=',')
                savetxt(dirli+'ASE.txt',Merge_ASE, fmt='%f', delimiter=',')
                savetxt(dirli+'AMI.txt',Merge_AMI, fmt='%f', delimiter=',')

                # savetxt(l0dir+'NodeStack'+'.txt',eva_NodeStack[0:Cutsum,:], fmt='%d', delimiter=',')
                # savetxt(l0dir+'ASE_aveg'+'.txt',eva_ASE_mean[0:Cutsum].T, fmt='%f', delimiter=',')
                savetxt(dirli+'MAX_corr.txt',Merge_Max_Similarity , fmt='%f', delimiter=',')
                savetxt(dirli+'Lamda.txt',Merge_Lamda, fmt='%f', delimiter=',')

                #savetxt(0ldir+'CE'+'.txt',eva_Ce[0:Cutsum], fmt='%f', delimiter=',')       
                logfilepointer.write("\nexp data researved")
        
        #result of  all layers
        savetxt(ldir+'EVANRMSE_test.txt',eva_NRMSE_test, fmt='%f', delimiter=',')
        savetxt(ldir+'EVANRMSE_train.txt',eva_NRMSE_train , fmt='%f', delimiter=',')
        savetxt(ldir+'EVAASE.txt',eva_ASE, fmt='%f', delimiter=',')   
        savetxt(ldir+'EVAAMI.txt',eva_AMI, fmt='%f', delimiter=',')   
        savetxt(ldir+'finalinc.txt',Inc_Trueout.T, fmt='%f', delimiter=',')  
        savetxt(ldir+'finaltrain.txt',ESN_0.Y_tt.T, fmt='%f', delimiter=',')       
        logfilepointer.write("\nevaluate_success")
        logfilepointer.close()
        return 
    
    
   
    @numba.jit
    def fastRepeatOptOram(self,Oram,redotime=1):
        #opt use InitX_dim
        self.Oram=Oram
        eva_NRMSE_train=zeros((redotime,))
        eva_NRMSE_test=zeros((redotime,))
        for r in range (redotime):
            # simulate ESN
            ESN_0 = EchoStateNetwork(Pram=self.Pram,Oram=self.Oram)
            #add nodes
            ESN_0.Inilize_First_reservoir(ESN_0.InitX_dim)
            for i in range(1 , ESN_0.Stacklayer):
                ESN_0.Inilize_Stack_a_reservoir(ESN_0.InitX_dim)
            
            ESN_0.Init_reservior(ESN_0.U_init)
            TrainProcessX=ESN_0.Train_reservoir(ESN_0.U_train,ESN_0.Y_train)
            
            ESN_0.Reinit_reservoir()
            #validate
            Yout,ValidateX=ESN_0.Validate_test_data_constant(ESN_0.U_tt)
            
            #estimate err
            err=(Yout-ESN_0.Y_tt)**2
            Ynorm=mean(ESN_0.Y_tt)*ones(Yout.shape)
            nerr=(Yout-Ynorm)**2
            #MSE_train=sum(err[0:ESN_0.TrainProcessLen])/ESN_0.TrainProcessLen
            NRMSE_train=sqrt(sum(err[0,0:ESN_0.TrainProcessLen])/sum(nerr[0,0:ESN_0.TrainProcessLen]))
            tlen=Yout.shape[1]-ESN_0.TrainProcessLen
            #MSE_test=sum(err[0,ESN_0.TrainProcessLen:])/tlen
            NRMSE_test=sqrt(sum(err[0,ESN_0.TrainProcessLen:Yout.shape[1]])/sum(nerr[0,ESN_0.TrainProcessLen:Yout.shape[1]]))
            eva_NRMSE_train[r]=NRMSE_train
            eva_NRMSE_test[r]=NRMSE_train

        #assignval
        NRMSE_train=mean(eva_NRMSE_train,axis=0)
        NRMSE_test=mean(eva_NRMSE_test,axis=0)
        return NRMSE_test+NRMSE_train
        return NRMSE_test
        return NRMSE_train
    def Repeat(self,cdir='./cache',redotime=1):
    
        self.mkdir(cdir)        
        #allocate storage
        eva_NRMSE_test=zeros((redotime,))
        eva_NRMSE_train=zeros(( redotime,))
        eva_ASE_mean=zeros(( redotime,))

        
        for r in range (redotime):
            ldir=cdir+str(r) +'/'        
            self.mkdir(ldir) 
            # simulate ESN
            ESN_0 = EchoStateNetwork(Pram=self.Pram,Oram=self.Oram)
            #add nodes
            ESN_0.Inilize_First_reservoir(ESN_0.InitX_dim)
            for i in range(1 , ESN_0.Stacklayer):
                ESN_0.Inilize_Stack_a_reservoir(ESN_0.InitX_dim)
            
            ESN_0.Init_reservior(ESN_0.U_init)
            TrainProcessX=ESN_0.Train_reservoir(ESN_0.U_train,ESN_0.Y_train)
            
            fig=plt.figure(1)
            fig.clear()
            axia=fig.gca()
            ces=axia.matshow(TrainProcessX)
            plt.colorbar(ces)
            plt.savefig(ldir+'X.eps')
            # plt.close(plt.figure(1))

            ESN_0.Reinit_reservoir()
            #validate
            Yout,ValidateX=ESN_0.Validate_test_data_constant(ESN_0.U_tt)
            #draw fig
            self.GUI_sig_and_err(Yout,ESN_0.Y_tt,ESN_0.TrainProcessLen,ldir,r)

            #estimate err
            err=(Yout-ESN_0.Y_tt)**2
            Ynorm=mean(ESN_0.Y_tt)*ones(Yout.shape)
            nerr=(Yout-Ynorm)**2
            NRMSE_train=sqrt(sum(err[0,0:ESN_0.TrainProcessLen])/sum(nerr[0,0:ESN_0.TrainProcessLen]))
            tlen=Yout.shape[1]-ESN_0.TrainProcessLen
            NRMSE_test=sqrt(sum(err[0,ESN_0.TrainProcessLen:Yout.shape[1]])/sum(nerr[0,ESN_0.TrainProcessLen:Yout.shape[1]]))
            eva_NRMSE_train[r]=NRMSE_train
            eva_NRMSE_test[r]=NRMSE_test
            if(Pre_def_USE_GPU): 
                X_num,T_num= ValidateX.shape    
                ASE=Entropy.EntropyEstimater(ValidateX,X_num,T_num)
            else:
                ASE=0
            eva_ASE_mean[r]=ASE
            #print(ASE)       
            print('r')
        #assignval
        self.DictRes[ 'NRMSEtrain_aveg']=mean(eva_NRMSE_train,axis=0)
        self.DictRes[ 'NRMSEtrain_std']=std(eva_NRMSE_train,axis=0)
        self.DictRes[ 'NRMSEtest_aveg']=mean(eva_NRMSE_test,axis=0)
        self.DictRes[ 'NRMSEtest_std']=std(eva_NRMSE_test,axis=0)
        self.DictRes[ 'EntropyX_aveg']=mean(eva_ASE_mean,axis=0)
        self.DictRes[ 'EntropyX_std']=std(eva_ASE_mean,axis=0)

        #draw NRMSE
        
        #savefiles
        if(1):
            savetxt(cdir+'REPNRMSE_test.txt',eva_NRMSE_test, fmt='%f', delimiter=',')
            savetxt(cdir+'REPNRMSE_train.txt',eva_NRMSE_train, fmt='%f', delimiter=',')
            savetxt(cdir+'REPASE.txt',eva_ASE_mean.T, fmt='%f', delimiter=',')       
        
        self.ExcelRecord(cdir+"pramAndResult")   

        return 
    def mkdir(self,path):
        # 引入os模块
        # 去除首位空格
        path=path.strip()
        # 去除尾部 \ 符号
        path=path.rstrip("\\")
        # 判断路径是否存在
        isExists=os.path.exists(path)    
        # 判断结果
        if not isExists:
            # 如果不存在则创建目录
            os.makedirs(path) 
            print (path+'创建成功')
            return True
        else:
            # 如果目录存在则不创建，并提示目录已存在
            print (path+' 目录已存在')
            return False
     
    def GUI_sig_and_err(self,Yo,Ys,TrainLen,dir,id:int):
        dim,ShowLen=shape(Yo)
        fig=plt.figure(1)
        fig.clear()
        for j in range(dim):
            axis=fig.add_subplot(dim,1,1+j)
            axis.plot(range(TrainLen,ShowLen), Ys[j:j+1,TrainLen:ShowLen].T, ls="-", lw=0.2,color='b', label='TestData')
            axis.plot(range(TrainLen), Ys[j:j+1,0:TrainLen].T,  ls="-",lw=0.2,color='b',label='TrainData')
            axis.plot(range(TrainLen), Yo[j:j+1,0:TrainLen].T, ls=":", lw=0.2,color='g',label='TrainData of Predicted')
            axis.plot(range(TrainLen,ShowLen), Yo[j:j+1,TrainLen:ShowLen].T, ls="-", lw=0.2,color='r', label='TestData of Predicted') 
        axis.legend(loc=0)
        fig.set_rasterized(True)
        plt.savefig(dir+'prediction_comparation'+str(id)+'.eps') 
    
        # plt.close(plt.figure(1))

        fig=plt.figure(2) 
        fig.clear()   
        for j in range(dim):
            axis=fig.add_subplot(dim,1,j+1)
            axis.plot(range(ShowLen), (Ys[j:j+1,0:ShowLen]-Yo[j:j+1,0:ShowLen]).T, ls="-", lw=0.2,color='r', label='Predicion error') 
        axis.legend(loc=0)
        fig.set_rasterized(True)
        plt.savefig(dir+'predicErr'+str(id)+'.eps')  
        # plt.close(plt.figure(1))

        return 

    def ExcelRecord(self,fname="./odata/def"):
        try:
            Wb= load_workbook(fname)
        except:
            Wb = Workbook()
        wp=Wb.active
        # ws.title='time'
        # t = datetime.datetime.now()
        # ws['A1']=t
        # values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10) # 选择图表的数据源
        # chart = BarChart() # 创建一个BarChart对象
        # chart.add_data(values) # 给BarChart对象添加数据源
        # ws.add_chart(chart, "E15") # 在工作表上添加图表，并指定图表左上角锚定的单元格。

        #page p
        wp = Wb.create_sheet()
        wp.append(['Prams','fixed'])
        # wp.append(list(self.Pram.keys()))
        # wp.append(list(self.Pram.values()))
        wp.append(['SuperPrams','optilized'])
        wp.append(list(self.Oram.keys()))
        wp.append(list(self.Oram.values()))
        #wp.append(['experiment','result'])
        wp.append(list(self.DictRes.keys()))
        wp.append(list(self.DictRes.values()))
        #page p done

        # Save the file
        try:
            Wb.save(fname+'.xlsx')
        except:
            Wb.save("./cache/def.xlsx")
        return
@numba.jit        
def  pso_func_wi(x):
    Opt_Pram=Default_Pram
    Opt_Oram=Default_Oram   
    Opt_Pram['Pram_Stack']=1 
    Opt_Oram['ampWi']=x[0]
    Es=ESNsimulater(Pram=Opt_Pram,Oram=Opt_Oram) 
 
    return  Es.fastRepeatOptOram(Opt_Oram,3)
@numba.jit  
def  pso_func_wp(x):
    Opt_Oram=Default_Oram
    Opt_Oram['ampWp']=x[0] 
    Es=ESNsimulater(Pram=Default_Pram,Oram=Opt_Oram) 
    # Opt_Oram['ampWi']=x[2] 
    # Opt_Oram['reg_fac']=x[3]*1e-3 
    return  Es.fastRepeatOptOram(Opt_Oram,3)

if __name__ == '__main__':
    Datesetdir='./ATP/'
    
    if(0 and Pre_def_PSO):
    #PSO wi
        Proceduledir=Datesetdir+'PSO/'+'MSO1/'
        pso = PSO(func=pso_func_wi, dim=1, pop=20, max_iter=20, lb=[0.0001], ub=[10], w=0.8, c1=0.5, c2=0.5)
        pso.run()
        opx=pso.gbest_x
        opres=pso.gbest_y
        print(opx)
        fig=plt.figure(1)
        fig.clear()
        axia=fig.gca()
        axia.plot(pso.gbest_y_hist)
        plt.savefig(Proceduledir+'psoof'+'.jpeg',)
        OptOram=Default_Oram   
        OptOram['ampWi']=opx[0]
        Es=ESNsimulater(Pram=Default_Pram,Oram=OptOram) 
        Es.Repeat(Proceduledir,5)


    if(0 and Pre_def_PSO):
        #PSO wp

        Proceduledir=Datesetdir+'PSO/'+'MSO1/'
        pso = PSO(func=pso_func_wp, dim=1, pop=20, max_iter=20, lb=[0.01], ub=[5], w=0.8, c1=0.5, c2=0.5)
        pso.run()
        opx=pso.gbest_x
        opres=pso.gbest_y
        print(opx)
        fig=plt.figure(1)
        fig.clear()
        axia=fig.gca()
        axia.plot(pso.gbest_y_hist)
        plt.savefig(Proceduledir+'psoof'+'.jpeg')
        OptOram=Default_Oram   
        OptOram['ampWp']=opx[0]
        Es=ESNsimulater(Pram=Default_Pram,Oram=OptOram) 
        Es.Repeat(Proceduledir,5)
    
    
        
    if(0):    
        #run repeat
        Proceduledir=Datesetdir+'REP/'+'exp0/'
        di=int(Default_Pram['Pram_X_add']/100*10)
        Rep_Pram=Default_Pram
        for i in range(10):
            Es=ESNsimulater(Rep_Pram,Default_Oram)
            Es.Repeat(Proceduledir+str(i)+'/',20)
            if(di<Rep_Pram['Pram_X_add']):
                Rep_Pram['Pram_X_add']=int(Rep_Pram['Pram_X_add']-di)
            else:
                print('error from REP:too less neurons')

    if(1):    
        #run revaluate  
    
        Es=ESNsimulater(Pram=Default_Pram,Oram=Default_Oram) 
        REP=3
        Proceduledir=Datesetdir+'EVA/'+'123/'
        for i in range(REP):
            Es.Increasingly_evaluateMergeHE(Proceduledir+'m0/'+str(i)+'/',method=0,Q2=0)
            Es.Increasingly_evaluateMergeHE(Proceduledir+'m1/'+str(i)+'/',method=1,Q2=0)
            # Es.CCN_evaluateMergeHE(Proceduledir+'m1/'+str(i)+'/',method=4,Q2=0)
            # Es.IPMA_evaluateMerge(Proceduledir+'m4/'+str(i)+'/',percentage=0.95,method=4,Q2=1)

    end=1